"""
_utils/_loader/xlsx.py
===========

.. module:: xlsx
   :platform: Unix
   :synopsis: read and write .xlsx-files

Module Overview
---------------

This module includes reader and writer for .xlsx-files.

Functions
---------
.. autofunction:: read_xlsx
.. autofunction:: write_xlsx

"""
import numpy as np
from openpyxl import load_workbook

from ._helper import to_cube
from ..._core import DataCube


def read_xlsx_to_dict(filepath: str) -> DataCube:
    data_list = []
    max_x = float('-inf')
    max_y = float('-inf')
    
    # Load the workbook and select the first sheet
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active  # Get the first sheet
    
    # Read headers
    headers = [cell.value for cell in sheet[1]]
    
    # Iterate over rows and create dictionaries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data_list.append(row_dict)
        
        # Update max_x and max_y
        x_value = row[headers.index('x')] if 'x' in headers else None
        y_value = row[headers.index('y')] if 'y' in headers else None
        
        if x_value is not None:
            x = max(max_x, x_value)
        if y_value is not None:
            y = max(max_y, y_value)
    
    cube = to_cube(data_list, len_x=x, len_y=y)
    
    return DataCube(cube, wavelengths=headers[2:])




def write_xlsx(datacube: np.array, wavelenghts: np.array, filename: str):
    """
    Write out a .xlsx file.

    :param datacube:
    :param wavelenghts:
    :param filename:
    :return:
    """
    shape = datacube.shape

    df = pd.DataFrame()

    cols = []

    for i in wavelenghts:
        cols.append(str(i))

    idx = []

    for y in range(shape[1]):
        for x in range(shape[0]):
            spec_ = datacube[x, y, :]

            df_tmp = pd.DataFrame(spec_).T

            df = df.append(df_tmp)

            idx.append(f'x:{x}; y:{y}')

    df.columns = cols

    df.insert(0, column='Point', value=idx)

    df = df.set_index('Point')

    df.to_excel(f'{filename}.xlsx')

